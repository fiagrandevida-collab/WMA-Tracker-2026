import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('data.xlsx')  # Replace with your path to the actual xlsx file

# Create a Dashboard sheet
if 'Dashboard' in workbook.sheetnames:
    dashboard = workbook['Dashboard']
else:
    dashboard = workbook.create_sheet(title='Dashboard')

# Create headers for the Dashboard
months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

# Create formulas to sum data from sheets AD, AH, AJ for each month
for i, month in enumerate(months, start=1):
    dashboard.cell(row=1, column=i, value=month)  # Headers
    dashboard.cell(row=2, column=i, value=f'=SUM(AD!{month}:AD!{month}) + SUM(AH!{month}:AH!{month}) + SUM(AJ!{month}:AJ!{month})')

# Save the modified workbook
workbook.save('data_with_dashboard.xlsx')  # Save under a new name or the same name as needed
